"""
Project Tracker API - FastAPI 应用主入口
提供完整的 REST API + 静态文件服务
"""

import os
import sys
import json
import re
from datetime import datetime, date
from io import BytesIO
from typing import Optional, List
from collections import defaultdict

# Ensure backend package is importable
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from fastapi import FastAPI, Depends, HTTPException, UploadFile, File, Query, Form
from fastapi.responses import FileResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from sqlalchemy.orm import Session
from sqlalchemy import func, extract, case

from database import get_db, db_session, engine
from models import Base, Customer, Project, SupplierQuote, PriceReference

# ── App Setup ──────────────────────────────────────────────

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
FRONTEND_DIR = os.path.join(BASE_DIR, 'frontend')
DATA_DIR = os.path.join(BASE_DIR, 'data')

os.makedirs(DATA_DIR, exist_ok=True)
os.makedirs(os.path.join(DATA_DIR, 'uploads'), exist_ok=True)

app = FastAPI(title='Project Tracker', version='1.0.0')

# ── 确保表已创建 ──────────────────────────────────────────
Base.metadata.create_all(bind=engine)


# ════════════════════════════════════════════════════════════
#  Customer APIs
# ════════════════════════════════════════════════════════════

@app.get('/api/customers')
def list_customers(search: str = '', db: Session = Depends(get_db)):
    q = db.query(Customer)
    if search:
        kw = f'%{search}%'
        q = q.filter(
            Customer.company_name.ilike(kw) |
            Customer.contact_name.ilike(kw) |
            Customer.phone.ilike(kw)
        )
    return q.order_by(Customer.updated_at.desc()).all()


@app.get('/api/customers/{customer_id}')
def get_customer(customer_id: int, db: Session = Depends(get_db)):
    c = db.query(Customer).get(customer_id)
    if not c:
        raise HTTPException(404, '客户不存在')
    return c


@app.post('/api/customers')
def create_customer(data: dict, db: Session = Depends(get_db)):
    c = Customer(
        company_name=data.get('company_name', ''),
        address=data.get('address', ''),
        contact_name=data.get('contact_name', ''),
        phone=data.get('phone', ''),
        email=data.get('email', ''),
    )
    db.add(c)
    db.commit()
    db.refresh(c)
    return c


@app.put('/api/customers/{customer_id}')
def update_customer(customer_id: int, data: dict, db: Session = Depends(get_db)):
    c = db.query(Customer).get(customer_id)
    if not c:
        raise HTTPException(404, '客户不存在')
    for field in ['company_name', 'address', 'contact_name', 'phone', 'email']:
        if field in data:
            setattr(c, field, data[field])
    db.commit()
    return c


@app.delete('/api/customers/{customer_id}')
def delete_customer(customer_id: int, db: Session = Depends(get_db)):
    c = db.query(Customer).get(customer_id)
    if not c:
        raise HTTPException(404, '客户不存在')
    db.delete(c)
    db.commit()
    return {'ok': True}


# ════════════════════════════════════════════════════════════
#  Project APIs
# ════════════════════════════════════════════════════════════

@app.get('/api/projects')
def list_projects(
    search: str = '',
    customer_id: Optional[int] = None,
    is_landed: Optional[bool] = None,
    db: Session = Depends(get_db)
):
    q = db.query(Project)
    if search:
        kw = f'%{search}%'
        q = q.filter(Project.name.ilike(kw) | Project.description.ilike(kw))
    if customer_id:
        q = q.filter(Project.customer_id == customer_id)
    if is_landed is not None:
        q = q.filter(Project.is_landed == is_landed)
    return q.order_by(Project.updated_at.desc()).all()


@app.get('/api/projects/{project_id}')
def get_project(project_id: int, db: Session = Depends(get_db)):
    p = db.query(Project).get(project_id)
    if not p:
        raise HTTPException(404, '项目不存在')
    return {
        **serialize(p),
        'customer': serialize(db.query(Customer).get(p.customer_id)),
        'supplier_quotes': [serialize(q) for q in p.supplier_quotes],
    }


@app.post('/api/projects')
def create_project(data: dict, db: Session = Depends(get_db)):
    p = Project(
        name=data.get('name', ''),
        description=data.get('description', ''),
        customer_id=data['customer_id'],
        quotation_date=parse_date(data.get('quotation_date')),
        final_price=data.get('final_price'),
        final_price_notes=data.get('final_price_notes', ''),
        final_margin=data.get('final_margin'),
        final_margin_notes=data.get('final_margin_notes', ''),
        cost_price=data.get('cost_price'),
        is_landed=data.get('is_landed', False),
        landed_date=parse_date(data.get('landed_date')),
        category=data.get('category', ''),
    )
    db.add(p)
    db.commit()
    db.refresh(p)

    # Handle supplier quotes
    quotes = data.get('supplier_quotes', [])
    for qd in quotes:
        q = SupplierQuote(
            project_id=p.id,
            supplier_company=qd.get('supplier_company', ''),
            contact_name=qd.get('contact_name', ''),
            phone=qd.get('phone', ''),
            email=qd.get('email', ''),
            product_service_detail=qd.get('product_service_detail', ''),
            price=qd.get('price'),
            currency=qd.get('currency', 'CNY'),
            category=qd.get('category', ''),
        )
        db.add(q)
    db.commit()

    # Update price references
    update_price_references(db)
    return get_project(p.id, db)


@app.put('/api/projects/{project_id}')
def update_project(project_id: int, data: dict, db: Session = Depends(get_db)):
    p = db.query(Project).get(project_id)
    if not p:
        raise HTTPException(404, '项目不存在')

    fields = ['name', 'description', 'customer_id', 'final_price', 'final_price_notes',
              'final_margin', 'final_margin_notes', 'cost_price', 'is_landed', 'category']
    for f in fields:
        if f in data:
            setattr(p, f, data[f])
    if 'quotation_date' in data:
        p.quotation_date = parse_date(data['quotation_date'])
    if 'landed_date' in data:
        p.landed_date = parse_date(data['landed_date'])

    # Replace supplier quotes if provided
    if 'supplier_quotes' in data:
        db.query(SupplierQuote).filter(SupplierQuote.project_id == project_id).delete()
        for qd in data['supplier_quotes']:
            q = SupplierQuote(
                project_id=p.id,
                supplier_company=qd.get('supplier_company', ''),
                contact_name=qd.get('contact_name', ''),
                phone=qd.get('phone', ''),
                email=qd.get('email', ''),
                product_service_detail=qd.get('product_service_detail', ''),
                price=qd.get('price'),
                currency=qd.get('currency', 'CNY'),
                category=qd.get('category', ''),
            )
            db.add(q)

    db.commit()
    update_price_references(db)
    return get_project(project_id, db)


@app.delete('/api/projects/{project_id}')
def delete_project(project_id: int, db: Session = Depends(get_db)):
    p = db.query(Project).get(project_id)
    if not p:
        raise HTTPException(404, '项目不存在')
    db.delete(p)
    db.commit()
    update_price_references(db)
    return {'ok': True}


# ════════════════════════════════════════════════════════════
#  Supplier Quote APIs
# ════════════════════════════════════════════════════════════

@app.get('/api/projects/{project_id}/quotes')
def list_quotes(project_id: int, db: Session = Depends(get_db)):
    return db.query(SupplierQuote).filter(
        SupplierQuote.project_id == project_id
    ).order_by(SupplierQuote.id).all()


# ════════════════════════════════════════════════════════════
#  Dashboard APIs
# ════════════════════════════════════════════════════════════

@app.get('/api/dashboard/summary')
def dashboard_summary(
    year: Optional[int] = None,
    period: str = 'month',  # month | quarter | year
    db: Session = Depends(get_db)
):
    """看板汇总：落地统计、毛利、金额"""
    q = db.query(Project)
    if year:
        q = q.filter(extract('year', Project.created_at) == year)

    projects = q.all()
    total_count = len(projects)
    landed_count = sum(1 for p in projects if p.is_landed)
    win_rate = round(landed_count / total_count * 100, 1) if total_count > 0 else 0
    total_final_price = sum(p.final_price or 0 for p in projects)
    total_margin = sum(p.final_margin or 0 for p in projects)

    # Group by period
    groups = defaultdict(lambda: {'count': 0, 'landed': 0, 'price': 0.0, 'margin': 0.0})

    for p in projects:
        d = p.quotation_date or p.created_at.date() if p.created_at else date.today()
        if period == 'month':
            key = f'{d.year}-{d.month:02d}'
        elif period == 'quarter':
            key = f'{d.year}-Q{(d.month - 1) // 3 + 1}'
        else:
            key = str(d.year)
        groups[key]['count'] += 1
        if p.is_landed:
            groups[key]['landed'] += 1
        groups[key]['price'] += (p.final_price or 0)
        groups[key]['margin'] += (p.final_margin or 0)

    sorted_groups = sorted(groups.items())

    return {
        'summary': {
            'total_count': total_count,
            'landed_count': landed_count,
            'win_rate': win_rate,
            'total_final_price': round(total_final_price, 2),
            'total_margin': round(total_margin, 2),
        },
        'groups': [
            {
                'period': k,
                'count': v['count'],
                'landed': v['landed'],
                'price': round(v['price'], 2),
                'margin': round(v['margin'], 2),
            }
            for k, v in sorted_groups
        ],
    }


# ════════════════════════════════════════════════════════════
#  Price Reference APIs
# ════════════════════════════════════════════════════════════

@app.get('/api/price-references')
def list_price_references(
    search: str = '',
    category: str = '',
    db: Session = Depends(get_db)
):
    q = db.query(PriceReference)
    if search:
        kw = f'%{search}%'
        q = q.filter(PriceReference.product_service_name.ilike(kw))
    if category:
        q = q.filter(PriceReference.category == category)
    return q.order_by(PriceReference.category, PriceReference.updated_at.desc()).all()


@app.get('/api/price-references/categories')
def list_price_categories(db: Session = Depends(get_db)):
    cats = db.query(PriceReference.category).distinct().order_by(PriceReference.category).all()
    return [c[0] for c in cats]


@app.delete('/api/price-references/{ref_id}')
def delete_price_reference(ref_id: int, db: Session = Depends(get_db)):
    """删除单条价格参考记录"""
    ref = db.query(PriceReference).get(ref_id)
    if not ref:
        raise HTTPException(404, '价格参考不存在')
    db.delete(ref)
    db.commit()
    return {'ok': True}


@app.post('/api/price-references/batch-delete')
def batch_delete_price_refs(data: dict, db: Session = Depends(get_db)):
    """批量删除价格参考"""
    ids = data.get('ids', [])
    if not ids:
        raise HTTPException(400, '请提供要删除的ID列表')
    deleted = db.query(PriceReference).filter(PriceReference.id.in_(ids)).delete(synchronize_session=False)
    db.commit()
    return {'ok': True, 'deleted': deleted}


def update_price_references(db: Session):
    """Rebuild price_references from all supplier_quotes"""
    db.query(PriceReference).delete()
    quotes = db.query(SupplierQuote).all()

    # Group by (category, product_service_detail)
    groups = defaultdict(list)
    for q in quotes:
        key = (q.category or '未分类', q.product_service_detail or '未指定')
        groups[key].append(q)

    for (cat, name), qlist in groups.items():
        prices = [q.price for q in qlist if q.price is not None]
        currencies = list(set(q.currency for q in qlist if q.currency))
        suppliers = list(set(q.supplier_company for q in qlist if q.supplier_company))
        latest = max((q.created_at for q in qlist if q.created_at), default=datetime.now())

        ref = PriceReference(
            category=cat,
            product_service_name=name,
            avg_price=round(sum(prices) / len(prices), 2) if prices else None,
            min_price=min(prices) if prices else None,
            max_price=max(prices) if prices else None,
            currency=currencies[0] if len(currencies) == 1 else 'MIXED',
            quote_count=len(qlist),
            supplier_list=json.dumps(suppliers, ensure_ascii=False),
            latest_quote_date=latest.date() if hasattr(latest, 'date') else latest,
        )
        db.add(ref)

    db.commit()


# ════════════════════════════════════════════════════════════
#  Import APIs (PDF / Excel)
# ════════════════════════════════════════════════════════════

@app.post('/api/import/quote')
async def import_quote(
    file: UploadFile = File(...),
    project_id: int = Form(None),
    db: Session = Depends(get_db)
):
    """导入供应商报价文件（PDF/Excel），自动解析并入库"""
    content = await file.read()
    ext = os.path.splitext(file.filename)[1].lower()

    if ext in ('.xlsx', '.xls'):
        parsed = parse_excel_quote(content)
    elif ext == '.pdf':
        parsed = parse_pdf_quote(content)
    elif ext == '.csv':
        parsed = parse_csv_quote(content)
    else:
        raise HTTPException(400, f'不支持的文件格式: {ext}，仅支持 .xlsx/.xls/.pdf/.csv')

    if not parsed:
        raise HTTPException(400, '未能从文件中解析出有效报价数据')

    saved_quotes = []
    for item in parsed:
        q = SupplierQuote(
            project_id=project_id,
            supplier_company=item.get('supplier_company', ''),
            contact_name=item.get('contact_name', ''),
            phone=item.get('phone', ''),
            email=item.get('email', ''),
            product_service_detail=item.get('product_service_detail', ''),
            price=item.get('price'),
            currency=item.get('currency', 'CNY'),
            category=item.get('category', ''),
        )
        db.add(q)
        db.flush()
        saved_quotes.append(serialize(q))

    db.commit()
    # 不自动更新价格参考，等待用户确认后手动入库

    return {
        'ok': True,
        'count': len(saved_quotes),
        'quotes': saved_quotes,
        'pending': True,  # 标记为待确认
    }


@app.post('/api/import/confirm')
def confirm_import(db: Session = Depends(get_db)):
    """确认入库：将已导入的报价加入价格参考"""
    update_price_references(db)
    return {'ok': True, 'message': '价格参考已更新'}


# ════════════════════════════════════════════════════════════
#  Quote Delete APIs
# ════════════════════════════════════════════════════════════

@app.delete('/api/quotes/{quote_id}')
def delete_quote(quote_id: int, db: Session = Depends(get_db)):
    """删除单条供应商报价"""
    q = db.query(SupplierQuote).get(quote_id)
    if not q:
        raise HTTPException(404, '报价不存在')
    db.delete(q)
    db.commit()
    update_price_references(db)
    return {'ok': True}


@app.post('/api/quotes/batch-delete')
def batch_delete_quotes(data: dict, db: Session = Depends(get_db)):
    """批量删除供应商报价"""
    ids = data.get('ids', [])
    if not ids:
        raise HTTPException(400, '请提供要删除的报价ID列表')
    deleted = db.query(SupplierQuote).filter(SupplierQuote.id.in_(ids)).delete(synchronize_session=False)
    db.commit()
    update_price_references(db)
    return {'ok': True, 'deleted': deleted}


def parse_excel_quote(content: bytes) -> List[dict]:
    """解析 Excel/BOQ 报价文件 — 支持多子表、工程量清单、层级结构"""
    import openpyxl
    wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
    results = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # ── 0. BOQ表数据集中在前列，限制读取范围提升性能 ──
        max_col = min(ws.max_column, 50)
        rows = list(ws.iter_rows(min_col=1, max_col=max_col, values_only=True))
        if not rows:
            continue

        # ── 1. 智能定位表头（前15行中找含关键词最多的行）──
        header_row_idx = 0
        best_score = 0
        for i in range(min(15, len(rows))):
            candidate = [str(c).strip().lower() if c is not None else '' for c in rows[i]]
            col_map_test = map_columns(candidate)
            score = len(col_map_test)
            if score > best_score:
                best_score = score
                header_row_idx = i
                col_map = col_map_test

        if best_score < 2:
            continue  # 需要至少识别到2个列

        # ── 2. 从表头检测币种 ──
        currency = 'CNY'
        header_row = rows[header_row_idx]
        for c in header_row:
            if c:
                s = str(c).upper()
                if 'MOP' in s or '澳門' in s or '澳门' in s:
                    currency = 'MOP'
                    break
                elif 'HKD' in s or '港幣' in s or '港币' in s or 'HK$' in s:
                    currency = 'HKD'
                    break

        # ── 3. 从文件名提取供应商信息 ──
        supplier_from_file = sheet_name.strip() if sheet_name else ''

        # ── 4. 提取数据行 ──
        current_category = ''
        for row in rows[header_row_idx + 1:]:
            # 跳过全空行
            cells_with_content = [c for c in row if c is not None and str(c).strip()]
            if not cells_with_content:
                continue

            # 提取第一个非空列作为编号/名称判断
            first_cell = str(row[0]).strip() if row[0] is not None and str(row[0]).strip() else ''
            second_cell = str(row[1]).strip() if len(row) > 1 and row[1] is not None and str(row[1]).strip() else ''

            # 检测节标题（只有第2列有内容，看起来像分类标题）
            is_section_header = False
            if not first_cell and second_cell and not any(
                str(row[i]).strip() if len(row) > i and row[i] else ''
                for i in range(2, len(row))
            ):
                section_text = second_cell
                if not any(c.isdigit() for c in section_text) and len(section_text) < 50:
                    is_section_header = True
                    current_category = section_text

            if is_section_header:
                continue

            # 跳过纯说明行（只有第二列有大量文字，无编号无价格）
            if not first_cell and len(cells_with_content) <= 2:
                continue

            # 跳过汇总行
            all_text = ' '.join(str(c) for c in row if c).lower()
            if any(kw in all_text for kw in ['合计', '總計', 'total', 'grand total', '小计']):
                continue

            item = extract_from_row_enhanced(row, col_map)

            # 应用检测到的币种和类别
            if not item.get('currency') or item.get('currency') == 'CNY':
                item['currency'] = currency
            if current_category and not item.get('category'):
                item['category'] = current_category
            if supplier_from_file and not item.get('supplier_company'):
                item['supplier_company'] = supplier_from_file

            # 只保留有价格 或 有明显产品名称的数据行
            if item.get('price') is not None or (
                item.get('product_service_detail') and
                len(item.get('product_service_detail', '')) > 3 and
                not item.get('product_service_detail', '').startswith('K.')
            ):
                results.append(item)

    return results


def parse_csv_quote(content: bytes) -> List[dict]:
    """解析 CSV 报价文件"""
    import io, csv
    content_str = content.decode('utf-8-sig')
    reader = csv.reader(io.StringIO(content_str))
    rows = list(reader)
    if not rows:
        return []

    header = [c.strip().lower() if c else '' for c in rows[0]]
    col_map = map_columns(header)
    results = []

    for row in rows[1:]:
        if not any(row):
            continue
        item = extract_from_row(row, col_map)
        if item.get('product_service_detail') or item.get('price'):
            results.append(item)

    return results


def parse_pdf_quote(content: bytes) -> List[dict]:
    """解析 PDF 报价文件"""
    import fitz
    doc = fitz.open(stream=content, filetype='pdf')
    full_text = ''
    for page in doc:
        full_text += page.get_text() + '\n'

    results = []
    # Try to extract tabular data from the text
    lines = [l.strip() for l in full_text.split('\n') if l.strip()]

    # Heuristic: try to find supplier info first
    supplier_company = ''
    contact_name = ''
    phone = ''
    email = ''

    for line in lines:
        # Look for supplier company name patterns
        if any(kw in line for kw in ['公司', '供应商', 'Supplier', 'Company']):
            supplier_company = line.split('：')[-1].split(':')[-1].strip() if supplier_company == '' else supplier_company
        # Email
        email_match = re.search(r'[\w.-]+@[\w.-]+\.\w+', line)
        if email_match and not email:
            email = email_match.group()
        # Phone
        phone_match = re.search(r'1[3-9]\d{9}|\d{3,4}-?\d{7,8}', line)
        if phone_match and not phone:
            phone = phone_match.group()

    # Try to find price info
    for line in lines:
        # Look for currency indicators
        currency = 'CNY'
        if 'HKD' in line or 'HK$' in line or '港币' in line or '港元' in line:
            currency = 'HKD'
        elif 'MOP' in line or 'MOP$' in line or '澳门' in line:
            currency = 'MOP'
        elif '$' in line and 'US' in line:
            currency = 'USD'

        # Try to extract price (number with decimal)
        prices = re.findall(r'(?:¥|￥|RMB|CNY|HKD|MOP|USD|\$)?\s*(\d{1,3}(?:,\d{3})*(?:\.\d{1,2})?)\s*(?:元|万)?', line)
        prices_clean = []
        for p in prices:
            try:
                val = float(p.replace(',', ''))
                if val > 0:
                    prices_clean.append(val)
            except:
                pass

        if prices_clean:
            # Assume the largest number is the price, and preceding text is the product
            max_price = max(prices_clean) if len(prices_clean) == 1 else prices_clean[-1]
            product = re.sub(r'(?:¥|￥|RMB|CNY|HKD|MOP|USD|\$)?\s*\d[\d,.]*\s*(?:元|万)?', '', line).strip()
            if product and len(product) >= 2:
                results.append({
                    'supplier_company': supplier_company,
                    'contact_name': contact_name,
                    'phone': phone,
                    'email': email,
                    'product_service_detail': product[:500],
                    'price': max_price,
                    'currency': currency,
                    'category': '',
                })

    # If no structured results found, return single entry with full text
    if not results:
        results = [{
            'supplier_company': supplier_company,
            'contact_name': contact_name,
            'phone': phone,
            'email': email,
            'product_service_detail': full_text[:1000],
            'price': None,
            'currency': 'CNY',
            'category': '',
        }]

    return results


def map_columns(header: List[str]) -> dict:
    """Map column names to standard field names (Simplified & Traditional Chinese)"""
    mapping = {
        'supplier_company': [
            '供应商', '供應商', '公司', 'supplier', 'company', '厂商', '廠商',
            '供货商', '供貨商', '卖方', '賣方', '单位名称', '單位名稱', '报价单位', '報價單位',
        ],
        'contact_name': [
            '联系人', '聯繫人', 'contact', '姓名', 'name', '业务员', '業務員', '销售', '銷售',
        ],
        'phone': [
            '电话', '電話', 'phone', 'tel', '手机', '手機', 'mobile', '联系方式', '聯繫方式', '传真', '傳真',
        ],
        'email': [
            '邮箱', '郵箱', 'email', 'mail', 'e-mail', '电子邮箱', '電子郵箱', '邮件', '郵件',
        ],
        'product_service_detail': [
            '产品', '產品', '服务', '服務', '描述', 'product', 'service', 'detail', 'description',
            '设备名称', '設備名稱', '型号', '型號', '品名', '项目', '項目', '名称', '名稱',
            '规格', '規格', '货物名称', '貨物名稱', '产品名称', '產品名稱', '产品型号', '產品型號',
            '规格型号', '規格型號', '规格参数', '規格參數', '物料名称', '物料名稱',
            '物料描述', '货品名称', '貨品名稱', '货物', '貨物', '内容', '內容',
            '说明', '說明', 'item', 'model', 'part', '物料', '项目名称', '項目名稱',
            '設備', '设备',
        ],
        'price': [
            '价格', '價格', '单价', '單價', '金额', '金額', 'price', 'amount', '报价', '報價',
            '总价', '總價', '小计', '小計', '含税单价', '含稅單價', '不含税单价',
            '含税总价', '含稅總價', '未税单价', '未稅單價', '未税总价', '未稅總價',
            '合计金额', '合計金額', '费用', '費用', '售价', '售價', '成本价', '成本價',
            'unit price', 'total', '小计金额', '小計金額', '总计', '總計',
        ],
        'currency': ['币种', '幣種', '货币', '貨幣', 'currency', '币别', '幣別', '货币单位', '貨幣單位'],
        'category': [
            '类别', '類別', '分类', '分類', 'category', 'type', '类型', '類型',
            '产品类别', '產品類別', '商品类别', '商品類別', '物料类别', '物料類別', '品类', '品類',
        ],
        'brand': ['品牌', '参考品牌', '參考品牌', '可接受品牌', '推荐品牌', '推薦品牌', 'brand', '指定品牌'],
    }

    result = {}
    for field, keywords in mapping.items():
        for i, h in enumerate(header):
            if any(kw in h for kw in keywords):
                result[field] = i
                break

    return result


def extract_from_row(row: tuple, col_map: dict) -> dict:
    """Extract fields from a row using column mapping (legacy)"""
    return extract_from_row_enhanced(row, col_map)


def extract_from_row_enhanced(row: tuple, col_map: dict) -> dict:
    """Enhanced extraction — handles Excel numeric types and multi-column detail"""
    item = {}
    for field, idx in col_map.items():
        if idx >= len(row):
            continue

        val = row[idx]

        # 特殊处理：product_service_detail 列可能为空，需检查相邻列
        if field == 'product_service_detail':
            s = str(val).strip() if val is not None else ''
            if not s or s == 'None':
                # 检查 idx+1, idx+2
                for offset in range(1, 4):
                    adj_idx = idx + offset
                    if adj_idx < len(row) and row[adj_idx] is not None:
                        adj_s = str(row[adj_idx]).strip()
                        if adj_s and adj_s != 'None' and len(adj_s) > 1:
                            item[field] = adj_s
                            break
                if field not in item:
                    item[field] = s if s else ''
            else:
                item[field] = s
            continue

        if val is None or (isinstance(val, str) and not val.strip()):
            continue
        if field == 'price':
            if isinstance(val, (int, float)):
                item[field] = float(val)
            else:
                try:
                    s = str(val).replace(',', '').replace('¥', '').replace('￥', '').replace(' ', '')
                    item[field] = float(s)
                except ValueError:
                    item[field] = None
        elif field == 'product_service_detail':
            existing = item.get(field, '')
            s = str(val).strip()
            if existing and s and s != 'None':
                item[field] = existing + ' ' + s
            else:
                item[field] = s
        else:
            s = str(val).strip()
            if s and s != 'None':
                # 币种不应是纯数字
                if field == 'currency' and re.match(r'^[\d.,]+$', s):
                    continue
                item[field] = s

    # ── 后处理：品牌信息追加到产品描述 ──
    brand_info = []
    for key in list(item.keys()):
        if key == 'brand' and item[key]:
            brand_info.append(item.pop(key))
    # 同时检查 brand 列的相邻列（如"可接受品牌"紧跟在"參考品牌"后）
    if 'brand' in col_map:
        brand_idx = col_map['brand']
        for offset in range(1, 3):
            adj_idx = brand_idx + offset
            if adj_idx < len(row) and row[adj_idx] is not None:
                adj_s = str(row[adj_idx]).strip()
                if adj_s and adj_s != 'None' and len(adj_s) > 1:
                    brand_info.append(adj_s)
                    break
    if brand_info:
        brands = ' | '.join(brand_info)
        pd = item.get('product_service_detail', '')
        if pd:
            item['product_service_detail'] = f"{pd} 【品牌: {brands}】"
        else:
            item['product_service_detail'] = f"【品牌: {brands}】"

    item.setdefault('supplier_company', '')
    item.setdefault('contact_name', '')
    item.setdefault('phone', '')
    item.setdefault('email', '')
    item.setdefault('product_service_detail', '')
    item.setdefault('price', None)
    item.setdefault('currency', 'CNY')
    item.setdefault('category', '')
    return item


# ════════════════════════════════════════════════════════════
#  Helper
# ════════════════════════════════════════════════════════════

def parse_date(val):
    if not val:
        return None
    if isinstance(val, date):
        return val
    try:
        return datetime.strptime(str(val)[:10], '%Y-%m-%d').date()
    except:
        return None


def serialize(obj):
    """Convert SQLAlchemy object to dict"""
    if obj is None:
        return None
    d = {}
    for c in obj.__table__.columns:
        val = getattr(obj, c.name)
        if isinstance(val, (datetime, date)):
            val = val.isoformat() if isinstance(val, datetime) else str(val)
        d[c.name] = val
    return d


# ── Static Files ────────────────────────────────────────────
if os.path.exists(os.path.join(FRONTEND_DIR, 'css')):
    app.mount('/css', StaticFiles(directory=os.path.join(FRONTEND_DIR, 'css')), name='css')
if os.path.exists(os.path.join(FRONTEND_DIR, 'js')):
    app.mount('/js', StaticFiles(directory=os.path.join(FRONTEND_DIR, 'js')), name='js')


@app.get('/')
def index():
    return FileResponse(os.path.join(FRONTEND_DIR, 'index.html'))


if __name__ == '__main__':
    import uvicorn
    uvicorn.run(app, host='0.0.0.0', port=8000)
