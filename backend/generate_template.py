#!/usr/bin/env python
"""生成导入模板 Excel 文件"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

def create_template(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "报价导入模板"

    # 表头（报价日期放在供应商信息后面）
    headers = [
        "供应商公司", "供应商联系人", "供应商电话", "供应商邮箱", "报价日期",
        "产品/服务名称", "设备型号", "数量", "单位",
        "单价", "币种", "产品类别", "品牌/备注"
    ]

    header_font = Font(name="Microsoft YaHei", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="185FA5", end_color="185FA5", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    data_font = Font(name="Microsoft YaHei", size=11)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    examples = [
        ["PROMISE COMPANY LIMITED", "Che Hoi Lei", "+853 28261055", "info@promise.com.mo", "2026-06-05",
         "紧急呼叫系统网关 Telligence Station Gateway", "NGGTWY2-HNGGWPWR-AK", 3, "unit",
         48220, "MOP", "呼叫/报警系统", "Ascom"],
        ["海康威视", "张三", "13800138000", "zhang@hikvision.com", "2026-05-20",
         "泊车识别摄像机 DS-TCP440-B", "DS-TCP440-B", 81, "unit",
         1810, "MOP", "安防摄像头", "Hikvision / 大华可选"],
        ["", "", "", "", "2026-04-15",
         "智能停车场管理终端机 DS-TPM400-FP", "DS-TPM400-FP", 3, "unit",
         10760, "MOP", "停车系统", ""],
    ]

    categories = [
        "安防摄像头", "门禁系统", "网络设备", "呼叫/报警系统",
        "停车系统", "软件/服务器", "显示设备", "机柜/配件",
        "线缆/管材", "电源设备", "广播/音响", "对讲系统", "其他设备"
    ]
    currencies = ["CNY", "HKD", "MOP"]

    for r, row_data in enumerate(examples, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = data_font
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = thin_border
        # 报价日期列格式
        date_cell = ws.cell(row=r, column=5)
        date_cell.number_format = 'YYYY-MM-DD'

    col_widths = [30, 15, 18, 28, 14, 45, 28, 8, 8, 15, 8, 18, 30]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # 币种下拉
    curr_val = DataValidation(type="list", formula1=f'"{",".join(currencies)}"', allow_blank=True)
    curr_val.error = "请选择 CNY / HKD / MOP"
    curr_val.errorTitle = "无效币种"
    ws.add_data_validation(curr_val)
    curr_val.add(f'K2:K1000')

    # 类别下拉
    cat_val = DataValidation(type="list", formula1=f'"{",".join(categories)}"', allow_blank=True)
    cat_val.error = "请从下拉列表选择产品类别"
    cat_val.errorTitle = "无效类别"
    ws.add_data_validation(cat_val)
    cat_val.add(f'L2:L1000')

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:M{len(examples) + 10}"

    wb.save(path)
    print(f"模板已生成: {path}")

if __name__ == '__main__':
    import sys, os
    out = sys.argv[1] if len(sys.argv) > 1 else os.path.join(os.path.dirname(__file__), '..', '..', 'frontend', 'template', 'import_template.xlsx')
    os.makedirs(os.path.dirname(out), exist_ok=True)
    create_template(out)
